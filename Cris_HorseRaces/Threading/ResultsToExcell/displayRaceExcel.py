import openpyxl
from openpyxl.styles import Font



def meetingResultsToExcell(queue, event):
    while not event.is_set() or not queue.empty():
        meeting = queue.get()
        fieldList = meeting[0]
        resultList = meeting[1]
        try:
            wb = openpyxl.load_workbook("Horse Race.xlsx")
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            wb.create_sheet("Race Results")
            wb.save('Horse Race.xlsx')
        sheet = wb.worksheets[0]
        sheet.title = "Race Fields"

    ### for field
        #insert race information into excel
        if sheet.max_row == 1:
            row = sheet.max_row
        else:
            row = sheet.max_row + 2
        sheet.merge_cells(f"A{row}:F{row}")
        cell = sheet[f"A{row}"]
        cell.font = Font(bold=True)
        sheet[f"A{row}"] = fieldList[0]
        fieldList.pop(0)

        #insert headers into excel

        headers = ("No", "Horse", "Br", "Rider", "Trainer", "Odds")
        row += 1

        sheet[f"A{row}"] = headers[0]
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"] = headers[1]
        sheet[f"B{row}"].font = Font(bold=True)
        sheet[f"C{row}"] = headers[2]
        sheet[f"C{row}"].font = Font(bold=True)
        sheet[f"D{row}"] = headers[3]
        sheet[f"D{row}"].font = Font(bold=True)
        sheet[f"E{row}"] = headers[4]
        sheet[f"E{row}"].font = Font(bold=True)
        sheet[f"F{row}"] = headers[5]
        sheet[f"F{row}"].font = Font(bold=True)


        for info in fieldList:
            row += 1
            sheet[f"A{row}"] = info[0]
            sheet[f"B{row}"] = info[1]
            sheet[f"C{row}"] = info[2]
            sheet[f"D{row}"] = info[3]
            sheet[f"E{row}"] = info[4]
            sheet[f"F{row}"] = info[5]
            wb.save('Horse Race.xlsx')

    ####for results
        sheet = wb["Race Results"]

        if sheet.max_row == 1:
            row = sheet.max_row
        else:
            row = sheet.max_row + 2
        sheet.merge_cells(f"A{row}:F{row}")
        cell = sheet[f"A{row}"]
        cell.font = Font(bold=True)
        sheet[f"A{row}"] = resultList[0]
        resultList.pop(0)

        headers = ("Placing", "StartPrice", "Br", "Rider", "Trainer")
        row += 1
        sheet[f"A{row}"] = headers[0]
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"] = headers[1]
        sheet[f"B{row}"].font = Font(bold=True)
        sheet[f"C{row}"] = headers[2]
        sheet[f"C{row}"].font = Font(bold=True)
        sheet[f"D{row}"] = headers[3]
        sheet[f"D{row}"].font = Font(bold=True)
        sheet[f"E{row}"] = headers[4]
        sheet[f"E{row}"].font = Font(bold=True)


        for info in resultList:
            row += 1
            sheet[f"A{row}"] = info[0]
            sheet[f"B{row}"] = info[1]
            sheet[f"C{row}"] = info[2]
            sheet[f"D{row}"] = info[3]
            sheet[f"E{row}"] = info[4]
            wb.save('Horse Race.xlsx')

